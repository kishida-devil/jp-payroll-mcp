"""地域別最低賃金を厚労省の一覧Excelから抽出する。

    curl -L -A "Mozilla/5.0" -o mw.xlsx https://www.mhlw.go.jp/content/11200000/001571219.xlsx
    python extract-minimum-wage.py

出力は src/data/minimum-wage.json にそのまま置き換えられる形。

年度は西暦の数値で持ち、元号は era_year に別途持つ。Excelの見出しは「令和７年度」
のような元号表記(しかも全角数字)なので、そのまま入れると年の比較ができなくなるし、
改元のたびに利用者側のコードが壊れる。ここで正規化しておくのが唯一の変換点。

--check を付けると、既存の src/data/minimum-wage.json と一致するかだけ確認して
書き込まない。改定前に「同じ入力から同じ出力が出るか」を確かめるための経路で、
これが通らないまま10月の改定を流すと、値の更新と同時に形式まで変わってしまう。
"""
import datetime
import json
import os
import sys

import openpyxl

EN = ["Hokkaido", "Aomori", "Iwate", "Miyagi", "Akita", "Yamagata", "Fukushima",
      "Ibaraki", "Tochigi", "Gunma", "Saitama", "Chiba", "Tokyo", "Kanagawa",
      "Niigata", "Toyama", "Ishikawa", "Fukui", "Yamanashi", "Nagano", "Gifu",
      "Shizuoka", "Aichi", "Mie", "Shiga", "Kyoto", "Osaka", "Hyogo", "Nara",
      "Wakayama", "Tottori", "Shimane", "Okayama", "Hiroshima", "Yamaguchi",
      "Tokushima", "Kagawa", "Ehime", "Kochi", "Fukuoka", "Saga", "Nagasaki",
      "Kumamoto", "Oita", "Miyazaki", "Kagoshima", "Okinawa"]

# 元号の初年に対応する西暦から1引いた値。平成N年 = 1988+N、令和N年 = 2018+N。
ERAS = {"平成": (1988, "H"), "令和": (2018, "R")}

ZEN = str.maketrans("０１２３４５６７８９", "0123456789")


def parse_fiscal_year(label: str):
    """「令和７年度」→ (2025, 'R7')。元号も全角数字も想定する。

    改元した年は「令和元年度」と書かれ、数字が入らない。西暦に直すと1年目なので
    元 = 1 として扱う。配信データも R1 で持っている。
    """
    s = str(label).translate(ZEN).strip()
    for name, (base, letter) in ERAS.items():
        if s.startswith(name):
            rest = s[len(name):]
            digits = "".join(ch for ch in rest if ch.isdigit())
            n = 1 if (not digits and rest.startswith("元")) else int(digits) if digits else None
            if n is None:
                raise ValueError(f"年度の数字が読めません: {label!r}")
            return base + n, f"{letter}{n}"
    raise ValueError(f"未知の元号です: {label!r}。ERAS に追加してください。")


def parse_date(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, (int, float)):
        # Excelのシリアル値。1900年うるう年バグのぶん基準を1899-12-30に取る。
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(v))).isoformat()
    return None


def extract(path="mw.xlsx"):
    ws = openpyxl.load_workbook(path, data_only=True)[
        openpyxl.load_workbook(path, data_only=True).sheetnames[0]]

    # 見出しは2列おき(金額・発効日の対)。シート名は改定のたびに変わるので使わない。
    years = []
    for c in range(2, ws.max_column + 1, 2):
        label = ws.cell(1, c).value
        if label:
            year, era = parse_fiscal_year(label)
            years.append((c, year, era))
    if not years:
        raise SystemExit("年度の見出しが1つも読めません。Excelの構造が変わった可能性があります。")

    prefectures = {}
    for i in range(47):
        row = 3 + i
        ja = str(ws.cell(row, 1).value or "").replace("　", "").strip()
        if not ja:
            raise SystemExit(f"{row}行目に都道府県名がありません。行の位置が変わった可能性があります。")
        history = []
        for c, year, era in years:
            amount = ws.cell(row, c).value
            if not isinstance(amount, (int, float)):
                continue
            history.append({
                "fiscal_year": year,
                "hourly_wage": int(amount),
                "effective_from": parse_date(ws.cell(row, c + 1).value),
                "era_year": era,
            })
        prefectures[EN[i]] = {"prefecture_ja": ja, "code": i + 1, "history": history}

    return {
        "latest_fiscal_year": years[-1][1],
        "prefectures": prefectures,
        "latest_era_year": years[-1][2],
    }


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    target = os.path.join(here, "..", "src", "data", "minimum-wage.json")

    data = extract()
    years = len(data["prefectures"]["Tokyo"]["history"])
    current = {k: v["history"][-1] for k, v in data["prefectures"].items()}
    lo = min(current.items(), key=lambda x: x[1]["hourly_wage"])
    hi = max(current.items(), key=lambda x: x[1]["hourly_wage"])

    print(f"fiscal years : {years}  (.. {data['latest_fiscal_year']} / {data['latest_era_year']})")
    print(f"prefectures  : {len(data['prefectures'])}")
    print(f"Tokyo latest : {current['Tokyo']}")
    print(f"lowest       : {lo[0]} {lo[1]['hourly_wage']}")
    print(f"highest      : {hi[0]} {hi[1]['hourly_wage']}")
    print(f"simple mean  : {round(sum(v['hourly_wage'] for v in current.values()) / 47, 1)}")

    encoded = json.dumps(data, ensure_ascii=False, separators=(",", ":"))

    if "--check" in sys.argv:
        with open(target, encoding="utf-8") as f:
            shipped = json.load(f)
        if shipped == data:
            print("\ncheck: 出力は現在配信中のデータと一致します。")
            return 0
        print("\ncheck: 出力が配信中のデータと一致しません。")
        if shipped.get("latest_fiscal_year") != data["latest_fiscal_year"]:
            print(f"  年度が違います: 配信中 {shipped.get('latest_fiscal_year')} / 抽出 {data['latest_fiscal_year']}")
            print("  → 新年度が公表されたのなら、--check なしで実行して取り込んでください。")
        else:
            print("  年度は同じなのに内容が違います。抽出ロジックか元データの変更を疑ってください。")
        return 1

    with open(target, "w", encoding="utf-8") as f:
        f.write(encoded)
    print(f"\nwrote {target} ({len(encoded.encode('utf-8'))} bytes)")
    print("次: npm test で検証し、npx wrangler deploy、その後 npm run rapidapi:prepare")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

import openpyxl, json, re
wb = openpyxl.load_workbook('r8.xlsx', data_only=True)
JIS = ["北海道","青森","岩手","宮城","秋田","山形","福島","茨城","栃木","群馬","埼玉","千葉","東京","神奈川","新潟","富山","石川","福井","山梨","長野","岐阜","静岡","愛知","三重","滋賀","京都","大阪","兵庫","奈良","和歌山","鳥取","島根","岡山","広島","山口","徳島","香川","愛媛","高知","福岡","佐賀","長崎","熊本","大分","宮崎","鹿児島","沖縄"]
EN = ["Hokkaido","Aomori","Iwate","Miyagi","Akita","Yamagata","Fukushima","Ibaraki","Tochigi","Gunma","Saitama","Chiba","Tokyo","Kanagawa","Niigata","Toyama","Ishikawa","Fukui","Yamanashi","Nagano","Gifu","Shizuoka","Aichi","Mie","Shiga","Kyoto","Osaka","Hyogo","Nara","Wakayama","Tottori","Shimane","Okayama","Hiroshima","Yamaguchi","Tokushima","Kagawa","Ehime","Kochi","Fukuoka","Saga","Nagasaki","Kumamoto","Oita","Miyazaki","Kagoshima","Okinawa"]

prefs, grades = {}, None
for i, sn in enumerate(wb.sheetnames):
    ws = wb[sn]
    h  = ws.cell(8, 6).value
    hl = ws.cell(8, 8).value
    prefs[EN[i]] = {
        "prefecture_ja": JIS[i],
        "code": i + 1,
        "health_insurance_rate": round(h, 5),
        "long_term_care_rate": round(hl - h, 5),
        "child_support_rate": round(ws.cell(8, 10).value, 5),
        "pension_rate": round(ws.cell(8, 12).value, 5),
    }
    if grades is None:
        grades = []
        for r in range(11, ws.max_row + 1):
            g, m = ws.cell(r, 1).value, ws.cell(r, 2).value
            if not isinstance(m, (int, float)) or g in (None, ''):
                continue
            mm = re.match(r'^\s*(\d+)\s*(?:\((\d+)\))?\s*$', str(g))
            if not mm:
                continue
            lo, hi = ws.cell(r, 3).value, ws.cell(r, 5).value
            grades.append({
                "health_grade": int(mm.group(1)),
                "pension_grade": int(mm.group(2)) if mm.group(2) else None,
                "standard_monthly_remuneration": int(m),
                "remuneration_from": int(lo) if isinstance(lo, (int, float)) else None,
                "remuneration_to":   int(hi) if isinstance(hi, (int, float)) else None,
            })

assert len(prefs) == 47 and len(grades) == 50, (len(prefs), len(grades))
# contiguity check: every yen value must map to exactly one grade
for a, b in zip(grades, grades[1:]):
    assert a["remuneration_to"] == b["remuneration_from"], (a, b)
pg = [g["pension_grade"] for g in grades if g["pension_grade"]]
print("prefectures:", len(prefs), "grades:", len(grades))
print("health grades:", grades[0]["health_grade"], "-", grades[-1]["health_grade"])
print("pension grades:", min(pg), "-", max(pg), "count", len(pg))
print("pension SMR range:", grades[3]["standard_monthly_remuneration"], "-",
      [g["standard_monthly_remuneration"] for g in grades if g["pension_grade"] == max(pg)][0])
print("boundaries contiguous: OK")

meta = {
  "fiscal_year": "R8", "gregorian_year": 2026,
  "effective_from": "2026-03", "source": "全国健康保険協会(協会けんぽ) 令和8年3月分からの保険料額表",
  "source_url": "https://www.kyoukaikenpo.or.jp/g7/cat330/sb3150/r08/r8ryougakuhyou3gatukara/",
  "license": "公共データ利用規約(第1.0版)",
  "child_care_contribution_rate": 0.0036,
  "bonus_cap_health_annual": 5730000,
  "bonus_cap_pension_monthly": 1500000,
  "rounding": "employee share: <=0.50 yen truncate, >0.50 yen round up (when employer deducts from salary)"
}
json.dump({"meta": meta, "prefectures": prefs, "grades": grades},
          open('kenpo_r8.json','w',encoding='utf-8'), ensure_ascii=False, separators=(',',':'))
import os; print("json bytes:", os.path.getsize('kenpo_r8.json'))
